"""
Stock Manager Pro — Streamlit Web Edition (Supabase Edition)
=============================================================
All data is saved permanently in Supabase cloud database.
Run with: streamlit run stock_app_streamlit.py
"""

import streamlit as st
import json, os, copy
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import io

# ── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Stock Manager Pro",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Custom CSS (dark theme) ───────────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #1e272e; }
    section[data-testid="stSidebar"] { background-color: #2f3640; }
    section[data-testid="stSidebar"] * { color: #ffffff !important; }
    [data-testid="metric-container"] {
        background-color: #2f3640;
        border: 1px solid #353b48;
        border-radius: 10px;
        padding: 12px;
    }
    [data-testid="metric-container"] label { color: #aaaaaa !important; }
    [data-testid="metric-container"] [data-testid="stMetricValue"] {
        color: #44bd32 !important; font-size: 1.6rem !important;
    }
    h1, h2, h3, h4 { color: #00a8ff !important; }
    p, label, span { color: #ecf0f1; }
    [data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
    .stButton > button {
        background-color: #273c75; color: white;
        border: none; border-radius: 6px; font-weight: bold;
    }
    .stButton > button:hover { background-color: #40739e; }
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div,
    .stDateInput > div > div > input {
        background-color: #353b48 !important;
        color: #ffffff !important;
        border: 1px solid #40739e !important;
    }
    .stTabs [data-baseweb="tab-list"] { background-color: #2f3640; border-radius: 8px; }
    .stTabs [data-baseweb="tab"] { color: #aaaaaa; }
    .stTabs [aria-selected="true"] { color: #00a8ff !important; font-weight: bold; }
    hr { border-color: #353b48; }
    .connection-ok  { padding:8px 14px; background:#1e8449; color:white; border-radius:6px; font-weight:bold; }
    .connection-err { padding:8px 14px; background:#c23616; color:white; border-radius:6px; font-weight:bold; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
LOW_STOCK = 10
SHIFTS    = ["Morning", "Evening", "Night"]

# ══════════════════════════════════════════════════════════════════════════════
# SUPABASE CONNECTION
# ══════════════════════════════════════════════════════════════════════════════

def get_supabase():
    """
    Returns a Supabase client using credentials stored in Streamlit secrets.
    Secrets are set in Streamlit Cloud → App Settings → Secrets.
    """
    try:
        from supabase import create_client
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception as e:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# DATA LAYER  — reads/writes a single JSON blob stored in Supabase
# (table: app_data, columns: id TEXT primary key, payload JSONB)
# ══════════════════════════════════════════════════════════════════════════════

ROW_ID = "stock_manager_main"   # single row that holds all app data

def _default_data():
    return {
        "items": {}, "history": [],
        "storage_stores": [], "storage_products": {},
        "storage_qty": {}, "storage_history": [],
        "stores_list": [], "store_products": {},
        "daily_shift_data": {}, "drop_log": [],
        "ledger_products": {}, "stock_transactions": [],
        "ledger_entries": [], "ledger_payments": [],
        "audit_log": [], "reconcile_log": [],
    }


@st.cache_data(ttl=15, show_spinner=False)   # cache for 15 s — auto-refresh
def _fetch_from_supabase():
    """Load the single JSON blob from Supabase. Cached for 15 seconds."""
    sb = get_supabase()
    if sb is None:
        return None
    try:
        res = sb.table("app_data").select("payload").eq("id", ROW_ID).execute()
        if res.data:
            return res.data[0]["payload"]
        # First run — insert default row
        d = _default_data()
        sb.table("app_data").insert({"id": ROW_ID, "payload": d}).execute()
        return d
    except Exception as e:
        st.error(f"Supabase read error: {e}")
        return None


def load_data():
    """Return data dict — from Supabase if connected, else local fallback."""
    d = _fetch_from_supabase()
    if d is None:
        # Fallback to local JSON while Supabase isn't configured yet
        if os.path.exists("stock_data.json"):
            with open("stock_data.json") as f:
                d = json.load(f)
        else:
            d = _default_data()
    # Ensure all keys exist
    for k, v in _default_data().items():
        d.setdefault(k, v)
    return d


def save_data(d):
    """Write data dict back to Supabase (upsert) and invalidate cache."""
    sb = get_supabase()
    if sb is None:
        # Fallback: save locally
        try:
            tmp = "stock_data.json.tmp"
            with open(tmp, "w") as f:
                json.dump(d, f, indent=2)
            os.replace(tmp, "stock_data.json")
        except Exception as e:
            st.error(f"Local save error: {e}")
        return
    try:
        sb.table("app_data").upsert({"id": ROW_ID, "payload": d}).execute()
        _fetch_from_supabase.clear()   # bust cache so next read is fresh
    except Exception as e:
        st.error(f"Supabase save error: {e}")


# ── Session state ─────────────────────────────────────────────────────────────
if "data" not in st.session_state:
    st.session_state.data = load_data()

data = st.session_state.data


def persist():
    """Save current data and refresh session."""
    save_data(st.session_state.data)
    st.session_state.data = load_data()


# ── Ledger helpers ────────────────────────────────────────────────────────────
def ledger_products():    return data.get("ledger_products", {})
def stock_transactions(): return data.get("stock_transactions", [])
def ledger_entries():     return data.get("ledger_entries", [])
def ledger_payments():    return data.get("ledger_payments", [])

def ledger_stock_balance(product):
    bal = 0
    for tx in stock_transactions():
        if tx["product"] == product:
            bal += int(tx.get("qty", 0) or 0) if tx["type"] == "receive" \
                   else -int(tx.get("qty", 0) or 0)
    return bal

def ledger_price_balance(product):
    bal = 0.0
    for e in ledger_entries():
        if e["product"] == product:
            bal += float(e.get("debit", 0) or 0)
            bal -= float(e.get("credit", 0) or 0)
    return bal


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📦 Stock Manager Pro")
    st.markdown("---")

    # Connection status badge
    sb = get_supabase()
    if sb:
        st.markdown('<span class="connection-ok">☁️ Supabase Connected</span>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<span class="connection-err">⚠️ Not Connected (local mode)</span>',
                    unsafe_allow_html=True)

    st.markdown("---")
    page = st.radio(
        "Navigation",
        ["🏠 Dashboard", "🏪 Storage Count", "📤 Drop Stock",
         "🕐 Shift Sale", "📒 Ledger", "⚙️ Setup Guide"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.caption(f"🕐 {datetime.now().strftime('%d %b %Y  %H:%M')}")
    if st.button("🔄 Refresh Data", use_container_width=True):
        _fetch_from_supabase.clear()
        st.session_state.data = load_data()
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# SETUP GUIDE PAGE  (shows when Supabase not yet configured)
# ══════════════════════════════════════════════════════════════════════════════
if page == "⚙️ Setup Guide":
    st.title("⚙️ Supabase Setup Guide")
    st.markdown("""
Follow these steps **once** to connect your app to Supabase so all data is saved permanently.

---

### Step 1 — Create a Free Supabase Account
1. Go to **[supabase.com](https://supabase.com)** → click **Start your project**
2. Sign up with GitHub or email
3. Click **New Project**
4. Fill in:
   - **Project name:** `stock-manager`
   - **Database password:** choose a strong password (save it!)
   - **Region:** choose nearest to you (e.g. Southeast Asia)
5. Click **Create new project** — wait ~2 minutes

---

### Step 2 — Create the Database Table
1. In your Supabase project, click **SQL Editor** (left sidebar)
2. Click **New Query**
3. Paste this SQL and click **Run ▶**:

```sql
CREATE TABLE app_data (
  id      TEXT PRIMARY KEY,
  payload JSONB NOT NULL DEFAULT '{}'
);
```

---

### Step 3 — Get Your API Keys
1. Go to **Project Settings** (gear icon, left sidebar)
2. Click **API**
3. Copy two things:
   - **Project URL** — looks like `https://xxxxxxxxxxx.supabase.co`
   - **anon public key** — a long string starting with `eyJ...`

---

### Step 4 — Add Keys to Streamlit Cloud
1. Go to **[share.streamlit.io](https://share.streamlit.io)**
2. Find your app → click **⋮ (3 dots)** → **Settings**
3. Click **Secrets** tab
4. Paste this (replace with YOUR actual values):

```toml
SUPABASE_URL = "https://your-project-id.supabase.co"
SUPABASE_KEY = "eyJ...your-anon-key..."
```

5. Click **Save** — your app will restart automatically

---

### Step 5 — Done! ✅
- The sidebar will show **☁️ Supabase Connected**
- All your data is now saved permanently in the cloud
- Works across all computers and browsers
""")

    st.divider()
    st.subheader("🔍 Test Your Connection")
    test_url = st.text_input("Paste your Supabase URL to test")
    test_key = st.text_input("Paste your Supabase anon key to test", type="password")
    if st.button("🔌 Test Connection"):
        if test_url and test_key:
            try:
                from supabase import create_client
                client = create_client(test_url, test_key)
                client.table("app_data").select("id").limit(1).execute()
                st.success("✅ Connection successful! Now add these keys to Streamlit Secrets.")
            except Exception as e:
                st.error(f"❌ Connection failed: {e}")
                st.info("Make sure you created the `app_data` table using the SQL in Step 2.")
        else:
            st.warning("Please enter both URL and Key.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🏠 Dashboard":
    st.title("📊 Dashboard")
    st.caption("Live summary of your stock across all stores")

    total_products  = len(ledger_products())
    total_stores    = len(data.get("storage_stores", []))
    drop_count      = len(data.get("drop_log", []))
    low_stock_items = [p for p in ledger_products()
                       if ledger_stock_balance(p) < LOW_STOCK]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 Total Products",   total_products)
    c2.metric("🏪 Stores",           total_stores)
    c3.metric("📤 Total Drops",      drop_count)
    c4.metric("⚠️ Low Stock Items",  len(low_stock_items),
              delta=f"-{len(low_stock_items)} below {LOW_STOCK}",
              delta_color="inverse")
    st.divider()

    st.subheader("📦 Product Stock Balances")
    if ledger_products():
        import pandas as pd
        rows = []
        for prod, meta in sorted(ledger_products().items()):
            bal   = ledger_stock_balance(prod)
            price = float(meta.get("price", 0))
            rows.append({
                "Product":          prod,
                "Unit":             meta.get("unit", "pcs"),
                "Stock Balance":    bal,
                "Price (₹)":        price,
                "Total Value (₹)":  round(bal * price, 2),
                "Status":           "⚠️ Low" if bal < LOW_STOCK else "✅ OK",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.info("No products yet. Go to 📒 Ledger → Products to add products.")

    st.divider()
    st.subheader("📤 Recent Drop Log (Last 10)")
    drop_log = sorted(data.get("drop_log", []),
                      key=lambda x: x.get("ts",""), reverse=True)[:10]
    if drop_log:
        import pandas as pd
        df2  = pd.DataFrame(drop_log)
        cols = [c for c in ["date","store","product","qty","note"] if c in df2.columns]
        st.dataframe(df2[cols], use_container_width=True, hide_index=True)
    else:
        st.info("No drops recorded yet.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 2 — STORAGE COUNT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🏪 Storage Count":
    st.title("🏪 Storage Count")
    tab1, tab2 = st.tabs(["📋 View / Edit Storage", "⚙️ Manage Stores & Products"])

    with tab1:
        storage_stores   = data.get("storage_stores", [])
        storage_products = data.get("storage_products", {})
        storage_qty      = data.get("storage_qty", {})

        if not storage_stores:
            st.warning("No stores added yet. Go to 'Manage Stores & Products' tab.")
        else:
            c1, c2 = st.columns(2)
            sel_store = c1.selectbox("Select Store", storage_stores)
            sel_date  = c2.date_input("Date", value=date.today())
            date_key  = sel_date.strftime("%Y-%m-%d")
            store_products = storage_products.get(sel_store, [])

            if not store_products:
                st.info(f"No products assigned to **{sel_store}**.")
            else:
                st.markdown(f"### 📅 {sel_store} — {date_key}")
                with st.form(key=f"storage_form_{sel_store}_{date_key}"):
                    header = st.columns([3,2,2,2,2,2,3])
                    for h, col in zip(["Product","MIH","RCV","DRP","EXC","NOTE","Balance"], header):
                        col.markdown(f"**{h}**")
                    def _sg_num(x):
                        try:    return float(x)
                        except: return 0.0
                    def _sg_fmt(v):
                        if str(v).strip() == "": return ""
                        f = _sg_num(v)
                        return str(int(f)) if f == int(f) else str(f)

                    new_vals = {}
                    new_mih  = {}
                    for prod in store_products:
                        pinfo    = storage_qty.get(sel_store,{}).get(prod,{})
                        existing = pinfo.get("daily",{}).get(date_key,{})
                        cols = st.columns([3,2,2,2,2,2,3])
                        cols[0].markdown(f"**{prod}**")
                        mih  = cols[1].text_input("", value=_sg_fmt(pinfo.get("must_in_hand","")), key=f"mih_{prod}_{date_key}", label_visibility="collapsed")
                        rcv  = cols[2].text_input("", value=_sg_fmt(existing.get("received","")),  key=f"rcv_{prod}_{date_key}", label_visibility="collapsed")
                        drp  = cols[3].text_input("", value=_sg_fmt(existing.get("dropped","")),   key=f"drp_{prod}_{date_key}", label_visibility="collapsed")
                        exc  = cols[4].text_input("", value=_sg_fmt(existing.get("exchange","")),  key=f"exc_{prod}_{date_key}", label_visibility="collapsed")
                        note = cols[5].text_input("", value=str(existing.get("note","")),          key=f"nt_{prod}_{date_key}",  label_visibility="collapsed")
                        bal_val = _sg_num(rcv) - _sg_num(drp) + _sg_num(exc)
                        bal_str = str(int(bal_val)) if bal_val == int(bal_val) else str(round(bal_val,2))
                        cols[6].markdown(f"**{bal_str}**")
                        new_vals[prod] = {"received":rcv, "dropped":drp, "exchange":exc, "note":note}
                        new_mih[prod]  = mih

                    if st.form_submit_button("💾 Save Storage Data", use_container_width=True):
                        sq = data.setdefault("storage_qty",{})
                        sq.setdefault(sel_store,{})
                        for prod, vals in new_vals.items():
                            pnode = sq[sel_store].setdefault(prod,{})
                            pnode.setdefault("daily",{})[date_key] = vals
                            if str(new_mih.get(prod,"")).strip() != "":
                                pnode["must_in_hand"] = new_mih[prod]
                        persist()
                        st.success(f"✅ Saved for {sel_store} on {date_key}")
                        st.rerun()

    with tab2:
        st.subheader("🏪 Add / Remove Stores")
        c1, c2 = st.columns([3,1])
        new_store = c1.text_input("New Store Name", placeholder="e.g. Main Warehouse")
        if c2.button("➕ Add Store", use_container_width=True):
            if new_store.strip():
                stores = data.setdefault("storage_stores",[])
                if new_store.strip() not in stores:
                    stores.append(new_store.strip())
                    persist(); st.success(f"Store '{new_store}' added!"); st.rerun()
                else:
                    st.warning("Store already exists.")

        stores = data.get("storage_stores",[])
        if stores:
            del_store = st.selectbox("Remove Store", ["-- select --"]+stores, key="del_store")
            if st.button("🗑 Delete Store", type="secondary"):
                if del_store != "-- select --":
                    data["storage_stores"].remove(del_store)
                    data.get("storage_products",{}).pop(del_store,None)
                    data.get("storage_qty",{}).pop(del_store,None)
                    persist(); st.success(f"'{del_store}' deleted."); st.rerun()

        st.divider()
        st.subheader("📦 Assign Products to Store")
        if stores:
            c1, c2 = st.columns(2)
            assign_store = c1.selectbox("Store", stores, key="assign_store")
            assign_prod  = c2.text_input("Product Name", placeholder="e.g. Milk 1L")
            if st.button("➕ Assign Product"):
                if assign_prod.strip():
                    sp = data.setdefault("storage_products",{})
                    sp.setdefault(assign_store,[])
                    if assign_prod.strip() not in sp[assign_store]:
                        sp[assign_store].append(assign_prod.strip())
                        persist(); st.success(f"'{assign_prod}' added to {assign_store}"); st.rerun()
                    else:
                        st.warning("Product already in this store.")
            sp = data.get("storage_products",{})
            if sp:
                st.markdown("**Current Assignments:**")
                for s, prods in sp.items():
                    if prods: st.markdown(f"🏪 **{s}**: {', '.join(prods)}")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 3 — DROP STOCK
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📤 Drop Stock":
    st.title("📤 Drop Stock")
    tab1, tab2 = st.tabs(["➕ New Drop", "📋 Drop Log"])

    with tab1:
        stores = data.get("storage_stores",[])
        if not stores:
            st.warning("No stores found. Add stores in Storage Count first.")
        else:
            c1, c2    = st.columns(2)
            drop_store = c1.selectbox("From Store", stores)
            drop_date  = c2.date_input("Date", value=date.today(), key="drop_date")
            store_prods = data.get("storage_products",{}).get(drop_store,[])
            if not store_prods:
                st.warning(f"No products assigned to **{drop_store}**.")
            else:
                drop_prod = st.selectbox("Product", store_prods)
                c1, c2    = st.columns(2)
                drop_qty  = c1.number_input("Quantity", min_value=1, value=1)
                drop_note = c2.text_input("Note (optional)")
                if st.button("📤 Record Drop", type="primary", use_container_width=True):
                    entry = {
                        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "date": drop_date.strftime("%Y-%m-%d"),
                        "store": drop_store, "product": drop_prod,
                        "qty": drop_qty, "note": drop_note,
                    }
                    data.setdefault("drop_log",[]).append(entry)
                    date_key = drop_date.strftime("%Y-%m-%d")
                    sq = data.setdefault("storage_qty",{})
                    sq.setdefault(drop_store,{}).setdefault(drop_prod,{}).setdefault("daily",{}).setdefault(date_key,{})
                    _cell = sq[drop_store][drop_prod]["daily"][date_key]
                    try:    existing_drp = float(_cell.get("dropped",0) or 0)
                    except: existing_drp = 0.0
                    _cell["dropped"] = str(existing_drp + drop_qty)
                    _cell.setdefault("received","")
                    _cell.setdefault("exchange","")
                    _cell.setdefault("note","")
                    persist(); st.success(f"✅ Dropped {drop_qty} × {drop_prod}"); st.rerun()

    with tab2:
        st.subheader("📋 All Drop Records")
        drop_log = sorted(data.get("drop_log",[]), key=lambda x: x.get("ts",""), reverse=True)
        if drop_log:
            import pandas as pd
            df   = pd.DataFrame(drop_log)
            cols = [c for c in ["date","store","product","qty","note","ts"] if c in df.columns]
            st.dataframe(df[cols], use_container_width=True, hide_index=True)
            if st.button("📥 Export Drop Log to Excel"):
                wb = Workbook(); ws = wb.active; ws.title = "Drop Log"
                hfil = PatternFill("solid",fgColor="273c75"); hfnt = Font(bold=True,color="FFFFFF")
                for ci,h in enumerate(["Date","Store","Product","Qty","Note","Timestamp"],1):
                    c=ws.cell(row=1,column=ci,value=h); c.fill=hfil; c.font=hfnt
                for row in drop_log:
                    ws.append([row.get("date",""),row.get("store",""),row.get("product",""),
                                row.get("qty",0),row.get("note",""),row.get("ts","")])
                buf=io.BytesIO(); wb.save(buf); buf.seek(0)
                st.download_button("⬇️ Download Excel", buf,
                    file_name=f"drop_log_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("No drops recorded yet.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 4 — SHIFT SALE
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🕐 Shift Sale":
    st.title("🕐 Shift Wise Sale")
    import pandas as pd
    stores = data.get("stores_list",[]) or data.get("storage_stores",[])
    if not stores:
        st.warning("No stores configured. Add stores in Storage Count first.")
    else:
        c1, c2    = st.columns(2)
        sel_store  = c1.selectbox("Store", stores)
        sel_date   = c2.date_input("Date", value=date.today(), key="shift_date")
        date_key   = sel_date.strftime("%Y-%m-%d")
        st.markdown(f"### 🏪 {sel_store} — {date_key}")

        shift_products = data.get("store_products",{}).get(sel_store,[])

        def _sn(x):
            try:    return float(x)
            except: return 0.0
        def _fmt(v):
            if str(v).strip() == "": return ""
            f = _sn(v)
            return str(int(f)) if f == int(f) else str(round(f,2))

        if not shift_products:
            st.info(f"No products assigned to **{sel_store}**. Assign products to this store first.")
        else:
            tab_m, tab_e, tab_n = st.tabs(["🌅 Morning","🌆 Evening","🌙 Night"])
            for shift, tab in zip(SHIFTS, [tab_m, tab_e, tab_n]):
                with tab:
                    shift_node = (data.get("daily_shift_data",{})
                                  .get(sel_store,{}).get(date_key,{}).get(shift,{}))
                    with st.form(key=f"shift_{sel_store}_{date_key}_{shift}"):
                        header = st.columns([4,2,2,2,2])
                        for h, col in zip(["Product","Drop","Start","End","Sale"], header):
                            col.markdown(f"**{h}**")
                        new_rows = {}
                        shift_total = 0.0
                        for prod in shift_products:
                            ex   = shift_node.get(prod,{})
                            cols = st.columns([4,2,2,2,2])
                            cols[0].markdown(f"**{prod}**")
                            dv = cols[1].text_input("", value=_fmt(ex.get("drop","")),  key=f"d_{shift}_{prod}_{date_key}", label_visibility="collapsed")
                            sv = cols[2].text_input("", value=_fmt(ex.get("start","")), key=f"s_{shift}_{prod}_{date_key}", label_visibility="collapsed")
                            ev = cols[3].text_input("", value=_fmt(ex.get("end","")),   key=f"e_{shift}_{prod}_{date_key}", label_visibility="collapsed")
                            any_filled = any(str(x).strip() != "" for x in (dv, sv, ev))
                            sale_val   = _sn(sv) + _sn(dv) - _sn(ev)
                            sale_str   = (str(int(sale_val)) if sale_val == int(sale_val) else str(round(sale_val,2))) if any_filled else ""
                            cols[4].markdown(f"**{sale_str}**")
                            shift_total += sale_val if any_filled else 0.0
                            new_rows[prod] = {"drop":dv, "start":sv, "end":ev,
                                              "sale": (sale_val if any_filled else "")}
                        st.markdown(f"**{shift} total sale:** {shift_total:g}")
                        if st.form_submit_button(f"💾 Save {shift} Shift", use_container_width=True):
                            dsd = data.setdefault("daily_shift_data",{})
                            dsd.setdefault(sel_store,{}).setdefault(date_key,{})[shift] = new_rows
                            persist(); st.success(f"✅ {shift} shift saved"); st.rerun()

            st.divider()
            st.subheader("📊 Daily Summary (sale per product, all shifts)")
            day_node = data.get("daily_shift_data",{}).get(sel_store,{}).get(date_key,{})
            summary_rows = []
            grand_total  = 0.0
            for prod in shift_products:
                row        = {"Product": prod}
                prod_total = 0.0
                has        = False
                for shift in SHIFTS:
                    sale = day_node.get(shift,{}).get(prod,{}).get("sale","")
                    row[shift] = _fmt(sale)
                    prod_total += _sn(sale)
                    if str(sale).strip() != "": has = True
                row["Total"] = (str(int(prod_total)) if prod_total == int(prod_total) else str(round(prod_total,2)))
                grand_total += prod_total
                if has:
                    summary_rows.append(row)
            if summary_rows:
                st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)
            else:
                st.info("No shift entries recorded for this store and date yet.")
            st.metric("Total Sale Today (all products × all shifts)", f"{grand_total:g}")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 5 — LEDGER
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📒 Ledger":
    st.title("📒 Stock Ledger")
    tab_recv, tab_issue, tab_pay, tab_prods, tab_export = st.tabs(
        ["📥 Receive Stock","📤 Issue Stock","💰 Payments","🛒 Products","📊 Export"])

    # ── Receive ──────────────────────────────────────────────────────────────
    with tab_recv:
        st.subheader("📥 Record Stock Received")
        prods = list(ledger_products().keys())
        if not prods:
            st.warning("No products yet. Go to 'Products' tab.")
        else:
            c1,c2       = st.columns(2)
            recv_prod   = c1.selectbox("Product", prods, key="recv_prod")
            recv_date   = c2.date_input("Date", value=date.today(), key="recv_date")
            c1,c2,c3    = st.columns(3)
            recv_qty    = c1.number_input("Quantity", min_value=1, value=1, key="recv_qty")
            recv_source = c2.text_input("Received From", key="recv_source")
            recv_price  = c3.number_input("Price per unit (₹)", min_value=0.0, value=0.0, key="recv_price")
            recv_notes  = st.text_input("Notes", key="recv_notes")
            if st.button("✅ Record Receive", type="primary", use_container_width=True):
                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                data.setdefault("stock_transactions",[]).append({
                    "ts":ts,"date":recv_date.strftime("%Y-%m-%d"),
                    "product":recv_prod,"type":"receive",
                    "qty":recv_qty,"source":recv_source,"notes":recv_notes,
                })
                total_val = round(recv_qty * recv_price, 2)
                prev_bal  = ledger_price_balance(recv_prod)
                data.setdefault("ledger_entries",[]).append({
                    "ts":ts,"date":recv_date.strftime("%Y-%m-%d"),
                    "product":recv_prod,"type":"receive",
                    "qty":recv_qty,"price":recv_price,
                    "debit":total_val,"credit":0.0,
                    "balance":round(prev_bal+total_val,2),
                    "source":recv_source,"notes":recv_notes,
                })
                persist(); st.success(f"✅ Received {recv_qty} × {recv_prod}"); st.rerun()

        st.divider(); st.subheader("Recent Receives")
        recv_txs = [t for t in stock_transactions() if t.get("type")=="receive"]
        if recv_txs:
            import pandas as pd
            df = pd.DataFrame(sorted(recv_txs,key=lambda x:x.get("ts",""),reverse=True)[:20])
            st.dataframe(df[[c for c in ["date","product","qty","source","notes"] if c in df.columns]],
                         use_container_width=True, hide_index=True)
        else: st.info("No receives recorded yet.")

    # ── Issue ─────────────────────────────────────────────────────────────────
    with tab_issue:
        st.subheader("📤 Record Stock Issued")
        prods = list(ledger_products().keys())
        if not prods:
            st.warning("No products yet.")
        else:
            c1,c2       = st.columns(2)
            issue_prod  = c1.selectbox("Product", prods, key="issue_prod")
            issue_date  = c2.date_input("Date", value=date.today(), key="issue_date")
            c1,c2       = st.columns(2)
            issue_qty   = c1.number_input("Quantity", min_value=1, value=1, key="issue_qty")
            issue_to    = c2.text_input("Issued To", key="issue_to")
            issue_notes = st.text_input("Notes", key="issue_notes")
            cur_bal = ledger_stock_balance(issue_prod)
            st.info(f"Current balance of **{issue_prod}**: **{cur_bal}** units")
            if st.button("✅ Record Issue", type="primary", use_container_width=True):
                if issue_qty > cur_bal:
                    st.error(f"❌ Not enough stock! Available: {cur_bal}")
                else:
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    data.setdefault("stock_transactions",[]).append({
                        "ts":ts,"date":issue_date.strftime("%Y-%m-%d"),
                        "product":issue_prod,"type":"issue",
                        "qty":issue_qty,"source":issue_to,"notes":issue_notes,
                    })
                    prod_price = float(ledger_products().get(issue_prod,{}).get("price",0))
                    total_val  = round(issue_qty * prod_price, 2)
                    prev_bal   = ledger_price_balance(issue_prod)
                    data.setdefault("ledger_entries",[]).append({
                        "ts":ts,"date":issue_date.strftime("%Y-%m-%d"),
                        "product":issue_prod,"type":"issue",
                        "qty":issue_qty,"price":prod_price,
                        "debit":0.0,"credit":total_val,
                        "balance":round(prev_bal-total_val,2),
                        "source":issue_to,"notes":issue_notes,
                    })
                    persist(); st.success(f"✅ Issued {issue_qty} × {issue_prod}"); st.rerun()

        st.divider(); st.subheader("Recent Issues")
        issue_txs = [t for t in stock_transactions() if t.get("type")=="issue"]
        if issue_txs:
            import pandas as pd
            df = pd.DataFrame(sorted(issue_txs,key=lambda x:x.get("ts",""),reverse=True)[:20])
            st.dataframe(df[[c for c in ["date","product","qty","source","notes"] if c in df.columns]],
                         use_container_width=True, hide_index=True)
        else: st.info("No issues recorded yet.")

    # ── Payments ──────────────────────────────────────────────────────────────
    with tab_pay:
        st.subheader("💰 Record Payment")
        prods = ["(General)"] + list(ledger_products().keys())
        c1,c2      = st.columns(2)
        pay_prod   = c1.selectbox("Product (or General)", prods)
        pay_date   = c2.date_input("Date", value=date.today(), key="pay_date")
        c1,c2      = st.columns(2)
        pay_amount = c1.number_input("Amount (₹)", min_value=0.0, value=0.0)
        pay_desc   = c2.text_input("Description / Paid To")
        pay_notes  = st.text_input("Notes", key="pay_notes")
        if st.button("✅ Record Payment", type="primary", use_container_width=True):
            if pay_amount > 0:
                ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                prod = pay_prod if pay_prod != "(General)" else ""
                data.setdefault("ledger_payments",[]).append({
                    "ts":ts,"date":pay_date.strftime("%Y-%m-%d"),
                    "product":prod,"amount":pay_amount,
                    "description":pay_desc,"notes":pay_notes,
                })
                if prod:
                    prev_bal = ledger_price_balance(prod)
                    data.setdefault("ledger_entries",[]).append({
                        "ts":ts,"date":pay_date.strftime("%Y-%m-%d"),
                        "product":prod,"type":"payment","qty":0,"price":0.0,
                        "debit":0.0,"credit":pay_amount,
                        "balance":round(prev_bal-pay_amount,2),
                        "source":pay_desc,"notes":pay_notes,
                    })
                persist(); st.success(f"✅ Payment ₹{pay_amount} recorded"); st.rerun()
            else: st.error("Amount must be > 0.")

        st.divider(); st.subheader("Payment History")
        pmts = sorted(data.get("ledger_payments",[]),key=lambda x:x.get("ts",""),reverse=True)
        if pmts:
            import pandas as pd
            df   = pd.DataFrame(pmts)
            cols = [c for c in ["date","product","description","amount","notes"] if c in df.columns]
            st.dataframe(df[cols], use_container_width=True, hide_index=True)
            st.metric("Total Payments", f"₹{sum(p.get('amount',0) for p in pmts):,.2f}")
        else: st.info("No payments yet.")

    # ── Products ──────────────────────────────────────────────────────────────
    with tab_prods:
        st.subheader("🛒 Manage Products")
        with st.form("add_product_form"):
            c1,c2,c3   = st.columns(3)
            prod_name  = c1.text_input("Product Name")
            prod_price = c2.number_input("Price per unit (₹)", min_value=0.0, value=0.0)
            prod_unit  = c3.selectbox("Unit", ["pcs","kg","g","L","mL","box","dozen","pack"])
            if st.form_submit_button("💾 Add / Update Product", use_container_width=True):
                if prod_name.strip():
                    data.setdefault("ledger_products",{})[prod_name.strip()] = {
                        "price":prod_price,"unit":prod_unit}
                    persist(); st.success(f"✅ '{prod_name}' saved!"); st.rerun()
                else: st.error("Product name cannot be empty.")

        st.divider(); st.subheader("Current Products")
        prods = ledger_products()
        if prods:
            import pandas as pd
            rows = [{"Product":k,"Price (₹)":v.get("price",0),
                     "Unit":v.get("unit","pcs"),
                     "Current Stock":ledger_stock_balance(k)}
                    for k,v in sorted(prods.items())]
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            del_prod = st.selectbox("Delete Product", ["-- select --"]+list(prods.keys()))
            if st.button("🗑 Delete Product", type="secondary"):
                if del_prod != "-- select --":
                    data["ledger_products"].pop(del_prod,None)
                    data["stock_transactions"] = [t for t in data.get("stock_transactions",[])
                                                  if t.get("product") != del_prod]
                    data["ledger_entries"] = [e for e in data.get("ledger_entries",[])
                                              if e.get("product") != del_prod]
                    persist(); st.success(f"🗑 Deleted: {del_prod}"); st.rerun()
        else: st.info("No products yet.")

    # ── Export ─────────────────────────────────────────────────────────────────
    with tab_export:
        st.subheader("📊 Export to Excel")
        if st.button("📥 Generate Excel", type="primary", use_container_width=True):
            wb   = Workbook()
            hfil = PatternFill("solid",fgColor="273c75")
            hfnt = Font(bold=True,color="FFFFFF",size=11)
            ctr  = Alignment(horizontal="center",vertical="center")
            grn  = PatternFill("solid",fgColor="EAFAF1")
            red  = PatternFill("solid",fgColor="FDEDEC")
            even = PatternFill("solid",fgColor="EBF5FB")
            odd  = PatternFill("solid",fgColor="D6EAF8")

            ws1 = wb.active; ws1.title = "Stock Ledger"
            for ci,h in enumerate(["Date","Product","Type","Qty","Source","Notes"],1):
                c=ws1.cell(row=1,column=ci,value=h); c.fill=hfil; c.font=hfnt; c.alignment=ctr
            for i,tx in enumerate(sorted(stock_transactions(),
                                         key=lambda x:(x.get("date",""),x.get("ts",""))),2):
                fill = grn if tx.get("type")=="receive" else red
                for ci,val in enumerate([tx.get("date",""),tx.get("product",""),
                                         tx.get("type","").capitalize(),tx.get("qty",""),
                                         tx.get("source",""),tx.get("notes","")],1):
                    c=ws1.cell(row=i,column=ci,value=val); c.fill=fill; c.alignment=ctr

            ws2 = wb.create_sheet("Product Summary")
            for ci,h in enumerate(["Product","Price (₹)","Unit","Stock Balance","Total Value (₹)"],1):
                c=ws2.cell(row=1,column=ci,value=h); c.fill=hfil; c.font=hfnt; c.alignment=ctr
            for i,(prod,meta) in enumerate(sorted(ledger_products().items()),2):
                bal=ledger_stock_balance(prod); price=float(meta.get("price",0))
                fill=even if i%2==0 else odd
                for ci,val in enumerate([prod,price,meta.get("unit","pcs"),bal,round(bal*price,2)],1):
                    c=ws2.cell(row=i,column=ci,value=val); c.fill=fill; c.alignment=ctr

            ws3 = wb.create_sheet("Payments")
            for ci,h in enumerate(["Date","Product","Description","Amount (₹)","Notes"],1):
                c=ws3.cell(row=1,column=ci,value=h); c.fill=hfil; c.font=hfnt; c.alignment=ctr
            for i,pmt in enumerate(sorted(ledger_payments(),
                                          key=lambda x:(x.get("date",""),x.get("ts",""))),2):
                fill=even if i%2==0 else odd
                for ci,val in enumerate([pmt.get("date",""),pmt.get("product","(General)"),
                                         pmt.get("description",""),pmt.get("amount",0),
                                         pmt.get("notes","")],1):
                    c=ws3.cell(row=i,column=ci,value=val); c.fill=fill; c.alignment=ctr

            ws4 = wb.create_sheet("Drop Log")
            for ci,h in enumerate(["Date","Store","Product","Qty","Note","Timestamp"],1):
                c=ws4.cell(row=1,column=ci,value=h); c.fill=hfil; c.font=hfnt; c.alignment=ctr
            for i,row in enumerate(sorted(data.get("drop_log",[]),key=lambda x:x.get("ts","")),2):
                fill=even if i%2==0 else odd
                for ci,val in enumerate([row.get("date",""),row.get("store",""),row.get("product",""),
                                         row.get("qty",""),row.get("note",""),row.get("ts","")],1):
                    c=ws4.cell(row=i,column=ci,value=val); c.fill=fill; c.alignment=ctr

            buf=io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("⬇️ Download Excel File", buf,
                file_name=f"stock_export_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
            st.success("✅ Excel ready! Click Download above.")
